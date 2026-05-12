"""Сидинг каталога SKU из реальных xlsx в переписки/files.

Идёт по всем 'ТЗ Приёмка/Отгрузка*.xlsx' и 'Опись для заливки*.xlsx',
выгребает уникальные тройки (barcode, article, name) и пишет в БД.

Запуск: python scripts/seed_catalog.py
"""
import os
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.config import REFERENCE_FILES_DIR  # noqa: E402
from src.db.session import db_session  # noqa: E402
from src.services.catalog_service import upsert_sku  # noqa: E402


# Эвристика по заголовкам колонок ТЗ Приёмка / Отгрузка / Описи.
# В разных файлах колонки немного разные — берём по индексам найденных названий.
TARGET_HEADERS = {
    "barcode": {"шк", "баркод", "баркод товара", "шк товара"},
    "article": {"артикул поставщика", "артикул товара", "артикул"},
    "name": {"название товара", "наименование", "название"},
}


def _norm(v) -> str:
    return str(v or "").strip().lower()


def _index_columns(headers: list[str]) -> dict[str, int]:
    idx = {}
    for i, h in enumerate(headers):
        for field, options in TARGET_HEADERS.items():
            if field in idx:
                continue
            if h in options:
                idx[field] = i
    return idx


def _harvest_file(path: Path) -> list[tuple[str, str, str]]:
    """Возвращает [(barcode, article, name), ...] из одного xlsx."""
    out = []
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    except Exception as e:
        print(f"  skip {path.name}: {e}")
        return out

    for sheet in wb.sheetnames:
        if sheet.lower() in {"операции", "указания", "инструкции"}:
            continue
        ws = wb[sheet]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue
        headers = [_norm(c) for c in header_row]
        idx = _index_columns(headers)
        if "barcode" not in idx:
            continue

        for r in rows_iter:
            if r is None:
                continue
            bc = r[idx["barcode"]] if idx["barcode"] < len(r) else None
            if bc is None or str(bc).strip() == "":
                continue
            bc = str(bc).strip()

            article = ""
            if "article" in idx and idx["article"] < len(r) and r[idx["article"]]:
                article = str(r[idx["article"]]).strip()
            article = article or bc

            name = ""
            if "name" in idx and idx["name"] < len(r) and r[idx["name"]]:
                name = str(r[idx["name"]]).strip()
            name = name or article

            out.append((bc, article, name))
    return out


def main() -> None:
    if not REFERENCE_FILES_DIR.exists():
        print(f"REFERENCE_FILES_DIR не найдена: {REFERENCE_FILES_DIR}")
        sys.exit(1)

    candidates: list[Path] = []
    for f in os.listdir(REFERENCE_FILES_DIR):
        if not (f.endswith(".xlsx") or f.endswith(".xls")):
            continue
        if f.endswith(".xls"):
            continue  # skip .xls — старый формат, не нужен для сидинга
        if f.startswith("ТЗ ") or f.startswith("ТЗ_") or "Опись для заливки" in f:
            candidates.append(REFERENCE_FILES_DIR / f)

    print(f"Найдено файлов для сидинга: {len(candidates)}")

    seen: dict[str, tuple[str, str]] = {}
    for path in candidates:
        print(f"→ {path.name}")
        for bc, article, name in _harvest_file(path):
            if bc not in seen:
                seen[bc] = (article, name)
            else:
                # Обновим имя если в существующей записи оно короче
                old_article, old_name = seen[bc]
                if len(name) > len(old_name):
                    seen[bc] = (old_article, name)

    print(f"\nУникальных SKU: {len(seen)}")

    created = 0
    existed = 0
    with db_session() as session:
        for bc, (article, name) in sorted(seen.items()):
            _, was_created = upsert_sku(
                session, barcode=bc, article=article, name=name
            )
            if was_created:
                created += 1
                print(f"  + {article} ({bc}) — {name[:40]}")
            else:
                existed += 1

    print(f"\nДобавлено новых: {created}, уже было: {existed}")


if __name__ == "__main__":
    main()
