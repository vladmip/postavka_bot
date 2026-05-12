from io import BytesIO

import openpyxl

from src.generators import generate_tz_priemka, generate_tz_otgruzka
from src.generators.tz_priemka import PriemkaRow
from src.generators.tz_otgruzka import OtgruzkaRow


def test_generate_priemka_basic():
    rows = [
        PriemkaRow(barcode="MILK-CHOCOLATE", name="Шоколад MrBeast", qty=300),
        PriemkaRow(barcode="3CHOC", name="Набор 3 шт", qty=100),
    ]
    data = generate_tz_priemka(rows)
    assert isinstance(data, bytes) and len(data) > 1000

    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb["Лист1"]
    assert ws.cell(1, 1).value == "ШК"
    assert ws.cell(2, 1).value == "MILK-CHOCOLATE"
    assert ws.cell(2, 2).value == "Шоколад MrBeast"
    assert ws.cell(2, 5).value == 300
    assert ws.cell(3, 1).value == "3CHOC"
    assert ws.cell(3, 5).value == 100


def test_generate_otgruzka_wb_only():
    rows = [
        OtgruzkaRow(barcode="MILK-CHOCOLATE", name="Шоколад", qty=300,
                    warehouse="Котовск: Питание", marketplace="wb"),
    ]
    data = generate_tz_otgruzka(rows)
    wb = openpyxl.load_workbook(BytesIO(data))
    assert "вб" in wb.sheetnames
    ws = wb["вб"]
    assert ws.cell(2, 1).value == "MILK-CHOCOLATE"
    assert ws.cell(2, 5).value == "Котовск: Питание"
    assert ws.cell(2, 8).value == 300


def test_generate_otgruzka_both_marketplaces():
    rows = [
        OtgruzkaRow(barcode="MILK-CHOCOLATE", name="Шоколад", qty=300,
                    warehouse="Котовск: Питание", marketplace="wb"),
        OtgruzkaRow(barcode="3CHOC", name="Набор", qty=50,
                    warehouse="МСК_МОЛЖАНИКОВО_3_ХАБ", marketplace="ozon"),
    ]
    data = generate_tz_otgruzka(rows)
    wb = openpyxl.load_workbook(BytesIO(data))
    assert "вб" in wb.sheetnames
    assert "озон" in wb.sheetnames

    ws_wb = wb["вб"]
    assert ws_wb.cell(2, 1).value == "MILK-CHOCOLATE"
    assert ws_wb.cell(2, 5).value == "Котовск: Питание"

    ws_oz = wb["озон"]
    assert ws_oz.cell(2, 1).value == "3CHOC"
    assert ws_oz.cell(2, 5).value == "МСК_МОЛЖАНИКОВО_3_ХАБ"
    assert ws_oz.cell(2, 8).value == 50
