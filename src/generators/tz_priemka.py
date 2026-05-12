"""Генератор 'ТЗ Приёмка' xlsx по шаблону ЛЕБЕР.

Шаблон: src/generators/templates/priemka_template.xlsx (лист 'Лист1', 9 колонок).
Колонки: ШК | Название товара | Цвет/размер | Поставщик | количество |
         примечание | Фото товара | Номер груза | Артикул Поставщика
"""
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import List, Optional

import openpyxl

TEMPLATE_PATH = Path(__file__).parent / "templates" / "priemka_template.xlsx"

DEFAULT_SUPPLIER = "ИП Баковец"

# Соответствие наших полей колонкам шаблона (1-based)
COLS = {
    "barcode": 1,
    "name": 2,
    "color_size": 3,
    "supplier": 4,
    "qty": 5,
    "note": 6,
    "photo": 7,
    "cargo_no": 8,
    "supplier_article": 9,
}


@dataclass
class PriemkaRow:
    barcode: str       # → колонка 'ШК' (фактически article)
    name: str
    qty: int
    supplier: str = DEFAULT_SUPPLIER
    color_size: str = ""
    note: str = ""
    photo: str = ""
    cargo_no: str = ""
    supplier_article: str = ""


def generate_tz_priemka(rows: List[PriemkaRow], out_path: Optional[Path] = None) -> bytes:
    """Рендерит xlsx и возвращает bytes. Если out_path передан — также сохраняет файл."""
    wb = openpyxl.load_workbook(str(TEMPLATE_PATH))
    ws = wb["Лист1"]

    for i, row in enumerate(rows, start=2):
        ws.cell(i, COLS["barcode"]).value = row.barcode
        ws.cell(i, COLS["name"]).value = row.name
        ws.cell(i, COLS["color_size"]).value = row.color_size or 0
        ws.cell(i, COLS["supplier"]).value = row.supplier
        ws.cell(i, COLS["qty"]).value = row.qty
        ws.cell(i, COLS["note"]).value = row.note or None
        ws.cell(i, COLS["photo"]).value = row.photo or None
        ws.cell(i, COLS["cargo_no"]).value = row.cargo_no or None
        ws.cell(i, COLS["supplier_article"]).value = row.supplier_article or row.barcode

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    if out_path is not None:
        out_path.write_bytes(data)

    return data
