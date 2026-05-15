"""Шаблон xlsx «Новая заявка»: широкая таблица для пакетного создания заявки.

Структура:
    артикул | название | <Кластер 1> | <Кластер 2> | ... | <Кластер N>
    OFFER1  | …        |             |             |     |
    OFFER2  | …        |             |             |     |

Юзер заполняет количества per кластер, шлёт обратно. Бот парсит широкий формат
через wide_ship_request parser и создаёт ОДНУ заявку с items по всем кластерам.

Артикулы и названия предзаполняются из user's OzonProduct catalog.
Кластеры берутся из Ozon API (cached cluster_list).
"""
from __future__ import annotations

import logging
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from sqlalchemy.orm import Session

from src.db.models import OzonProduct

logger = logging.getLogger("generators.new_request_template")


def generate_template(
    session: Session,
    user_id: int,
    cluster_names: List[str],
) -> bytes:
    products = (
        session.query(OzonProduct)
        .filter(OzonProduct.user_id == user_id)
        .order_by(OzonProduct.offer_id)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявка"

    fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
    thin = Side(border_style="thin", color="FF999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True, name="Arial", size=11)
    normal = Font(name="Arial", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Формат как в Ozon ЛК экспорте (основной.xlsx):
    # R1: A1=пусто, B1+=имена кластеров.
    # R2..: A=offer_id, B+=количество per cluster.
    headers = [""] + cluster_names
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = bold
        c.fill = fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "B2"

    for row_idx, p in enumerate(products, start=2):
        ws.cell(row=row_idx, column=1, value=p.offer_id).alignment = left
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = border
            if cell.font is None or cell.font.name != "Arial":
                cell.font = normal
            if c > 1:
                cell.alignment = center

    ws.column_dimensions["A"].width = 28
    from openpyxl.utils import get_column_letter
    for i, _ in enumerate(cluster_names):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
