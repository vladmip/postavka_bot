"""Генератор 'ТЗ Отгрузка' xlsx по шаблону ЛЕБЕР.

Шаблон: src/generators/templates/otgruzka_template.xlsx (лист 'Лист1', 14 колонок).
Колонки: ШК | Название товара | Цвет/размер | Поставщик | Склад назначения |
         Транзитный склад | Тип отгрузки | количество | количество паллет |
         количество в коробке | Упаковка для товара | примечание |
         Фото товара | Артикул Поставщика

В реальных ТЗ Отгрузка пользователь использует 2 листа: 'вб' и 'озон'.
Этот генератор создаёт листы по факту наличия позиций для каждого МП:
  только wb     → лист 'вб'
  только ozon   → лист 'озон'
  оба           → 'вб' + 'озон'
"""
from copy import copy
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import List, Optional

import openpyxl

TEMPLATE_PATH = Path(__file__).parent / "templates" / "otgruzka_template.xlsx"

DEFAULT_SUPPLIER = "ИП Баковец"

COLS = {
    "barcode": 1,
    "name": 2,
    "color_size": 3,
    "supplier": 4,
    "warehouse": 5,
    "transit_warehouse": 6,
    "shipment_type": 7,
    "qty": 8,
    "pallets": 9,
    "qty_in_box": 10,
    "packaging": 11,
    "note": 12,
    "photo": 13,
    "supplier_article": 14,
}


@dataclass
class OtgruzkaRow:
    barcode: str
    name: str
    qty: int
    warehouse: str
    marketplace: str  # 'wb' | 'ozon'
    supplier: str = DEFAULT_SUPPLIER
    color_size: str = ""
    transit_warehouse: str = ""
    shipment_type: str = ""  # 'короба' | 'паллеты'
    pallets: Optional[int] = None
    qty_in_box: Optional[int] = None
    packaging: str = ""
    note: str = ""
    photo: str = ""
    supplier_article: str = ""


def generate_tz_otgruzka(rows: List[OtgruzkaRow], out_path: Optional[Path] = None) -> bytes:
    """Рендерит xlsx с листами 'вб'/'озон' в зависимости от состава позиций."""
    wb_wb = [r for r in rows if r.marketplace == "wb"]
    rows_oz = [r for r in rows if r.marketplace == "ozon"]

    book = openpyxl.load_workbook(str(TEMPLATE_PATH))
    base = book["Лист1"]

    if wb_wb and rows_oz:
        base.title = "вб"
        _fill_sheet(base, wb_wb)
        ozon_sheet = book.copy_worksheet(base)
        ozon_sheet.title = "озон"
        _clear_data_rows(ozon_sheet)
        _fill_sheet(ozon_sheet, rows_oz)
    elif wb_wb:
        base.title = "вб"
        _fill_sheet(base, wb_wb)
    elif rows_oz:
        base.title = "озон"
        _fill_sheet(base, rows_oz)
    else:
        # пустая поставка — оставляем шаблон как есть
        pass

    buf = BytesIO()
    book.save(buf)
    data = buf.getvalue()

    if out_path is not None:
        out_path.write_bytes(data)

    return data


def _fill_sheet(ws, rows: List[OtgruzkaRow]) -> None:
    for i, row in enumerate(rows, start=2):
        ws.cell(i, COLS["barcode"]).value = row.barcode
        ws.cell(i, COLS["name"]).value = row.name
        ws.cell(i, COLS["color_size"]).value = row.color_size or 0
        ws.cell(i, COLS["supplier"]).value = row.supplier
        ws.cell(i, COLS["warehouse"]).value = row.warehouse
        ws.cell(i, COLS["transit_warehouse"]).value = row.transit_warehouse or None
        ws.cell(i, COLS["shipment_type"]).value = row.shipment_type or None
        ws.cell(i, COLS["qty"]).value = row.qty
        ws.cell(i, COLS["pallets"]).value = row.pallets if row.pallets is not None else None
        ws.cell(i, COLS["qty_in_box"]).value = row.qty_in_box if row.qty_in_box is not None else None
        ws.cell(i, COLS["packaging"]).value = row.packaging or None
        ws.cell(i, COLS["note"]).value = row.note or None
        ws.cell(i, COLS["photo"]).value = row.photo or None
        ws.cell(i, COLS["supplier_article"]).value = row.supplier_article or row.barcode


def _clear_data_rows(ws) -> None:
    """Очистить данные начиная со 2-й строки, оставив заголовок."""
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None
