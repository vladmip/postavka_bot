"""Шаблон xlsx со всеми Ozon-артикулами для заполнения упаковки/примечаний.

Структура листа:
    R1: артикул | название | упаковка | примечание
    R2+: одна строка на товар из ozon_products, упаковка/примечание подставлены
         из ProductHint если есть (иначе пусто).

Колонка «название» — справочная (бот её не читает), нужна юзеру чтобы понять
о каком товаре речь. Парсер ищет колонки по подстроке, лишние игнорируются.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from sqlalchemy.orm import Session

from src.db.models import OzonProduct, ProductHint


def generate_hints_template(session: Session) -> bytes:
    products = session.query(OzonProduct).order_by(OzonProduct.offer_id).all()
    hints = {h.ozon_product_id: h for h in session.query(ProductHint).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Упаковка"

    header_fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
    thin = Side(border_style="thin", color="FF999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True, name="Arial", size=11)
    normal = Font(name="Arial", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["артикул", "название", "упаковка", "примечание"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[1].height = 28

    for row_idx, p in enumerate(products, start=2):
        h = hints.get(p.id)
        ws.cell(row=row_idx, column=1, value=p.offer_id).alignment = left
        ws.cell(row=row_idx, column=2, value=p.name or "").alignment = left
        ws.cell(row=row_idx, column=3, value=(h.packaging if h else "") or "").alignment = left
        ws.cell(row=row_idx, column=4, value=(h.notes if h else "") or "").alignment = left
        for c in range(1, 5):
            cell = ws.cell(row=row_idx, column=c)
            cell.border = border
            if cell.font is None or cell.font.name != "Arial":
                cell.font = normal

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 40
    ws.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
