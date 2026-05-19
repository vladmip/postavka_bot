"""Генератор ТЗ Отгрузка xlsx для ShipmentRequest.

Структура соответствует эталону пользователя (`файлы для показа клоду/ТЗ Отгрузка
шаблон.xlsx`):

  Лист 'вб':
    R1: 'ШК' | 'Название товара' | 'Поставщик' | <колонки складов/кластеров…>
        | 'упаковка' | 'состав набора' | 'Артикул Поставщика'
    R2: <empty> | <empty> | <empty> | <supply_id> | …          (по факту бронирования)
    R3: <empty> | <empty> | <empty> | <дата+таймслот> | …      (по факту бронирования)
    R4+: данные (одна строка на SKU, qty в нужный warehouse-столбец)

  Лист 'озон':
    R1: 'ШК' | 'Название товара' | 'Поставщик' | <колонки кластеров…>
        | 'Упаковка для товара' | 'примечание' | 'Артикул Поставщика'
        | 'Полный адрес Озон склада и таймслот'
    R2+: данные. Полное имя кластера в шапке (не сокращаем). Колонка K заполняется
         склад+таймслотом после бронирования.

  Лист 'операции': metadata (заявка, даты) + место для ручных заметок (merged A:B).

Колонки строятся динамически по фактическим направлениям заявки. Если у позиции
известен target_warehouse / booked_supply_id / booked_slot_at — генератор
выделяет для этого warehouse отдельную колонку и заполняет R2/R3. Иначе колонка
объединена по кластеру и R2/R3 пустые (заполнятся вручную после бронирования).
"""
from __future__ import annotations

import logging
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import object_session

from src.db.models import ShipmentRequest, ProductHint

logger = logging.getLogger("generators.ship_tz")

DEFAULT_SUPPLIER = "ИП Баковец"

_MONTHS_RU = [
    "", "янв", "фев", "мар", "апр", "мая", "июн",
    "июл", "авг", "сен", "окт", "ноя", "дек",
]


def _header_fill() -> PatternFill:
    return PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _thin_border() -> Border:
    s = Side(style="thin", color="999999")
    return Border(left=s, right=s, top=s, bottom=s)


def _fmt_slot(slot_at) -> str:
    """`datetime(2026,5,7,18,0)` → `'7 мая 18-19'`."""
    if not slot_at:
        return ""
    h_from = slot_at.hour
    h_to = (h_from + 1) % 24
    return f"{slot_at.day} {_MONTHS_RU[slot_at.month]} {h_from}-{h_to}"


# Уникальная «колонка» в листе ТЗ — по кластеру и (если есть) по конкретному
# складу бронирования. Если warehouse известен — будет отдельная колонка
# с supply_id и таймслотом в шапке. Иначе — общая колонка кластера, R2/R3 пусто.
ColKey = Tuple[str, Optional[str], Optional[str], Optional[str]]
# (cluster, target_warehouse_or_None, supply_id_or_None, slot_str_or_None)


def _clean_warehouse_name(wh: Optional[str]) -> Optional[str]:
    """Отфильтровать «псевдо-имена» типа '#0' / '#123' для CROSSDOCK-поставок,
    где Ozon не возвращает реального РФЦ. На таких item'ах в ТЗ Отгрузка
    колонкой должен быть кластер, а не '#0'."""
    if not wh:
        return None
    s = wh.strip()
    if not s:
        return None
    if s.startswith("#") and s[1:].isdigit():
        return None
    return s


def _build_columns(req: ShipmentRequest, marketplace: str) -> List[ColKey]:
    """Список уникальных колонок для листа, в порядке появления."""
    seen: List[ColKey] = []
    for it in req.items:
        if it.marketplace != marketplace:
            continue
        slot = _fmt_slot(it.booked_slot_at) if it.booked_slot_at else None
        key: ColKey = (
            it.cluster,
            _clean_warehouse_name(it.target_warehouse),
            it.booked_supply_id or None,
            slot,
        )
        if key not in seen:
            seen.append(key)
    return seen


def _col_label(col: ColKey) -> str:
    """Шапка колонки: имя склада если есть, иначе кластер."""
    cluster, wh, _supply, _slot = col
    return wh or cluster


def generate_ship_tz(req: ShipmentRequest) -> bytes:
    """Сгенерировать ТЗ Отгрузка xlsx для заявки. Возвращает bytes.

    Источник данных: ozon_products / wb_products (через item.ozon_product / item.wb_product).
    Ключ группировки SKU — внутренний id маркет-каталога (ozon_product.id или wb_product.id),
    одинаковый сквозной для уникальной товарной строки.

    Упаковка/примечание для Ozon-строк подтягиваются из ProductHint по ozon_product_id.
    """
    by_col_sku: Dict[Tuple[ColKey, int], int] = {}
    skus_meta: Dict[int, Dict] = {}  # product_id → {barcode, name, article, hint_packaging, hint_notes}
    unmatched: List[Dict] = []

    # Hints для Ozon-товаров заявки.
    ozon_hints: Dict[int, ProductHint] = {}
    session = object_session(req)
    if session is not None:
        oz_pids = [it.ozon_product_id for it in req.items if it.ozon_product_id]
        if oz_pids:
            rows = session.query(ProductHint).filter(
                ProductHint.ozon_product_id.in_(set(oz_pids))
            ).all()
            ozon_hints = {h.ozon_product_id: h for h in rows}

    def _product_key_and_meta(it):
        """Вернуть (key_id, meta_dict) для item или None если не привязан."""
        mp = (it.marketplace or "").lower()
        if mp == "ozon" and it.ozon_product:
            p = it.ozon_product
            h = ozon_hints.get(p.id)
            return p.id, {
                "barcode": p.barcode_primary or "",
                "name": p.name or "",
                "article": p.offer_id,
                "hint_packaging": (h.packaging if h else None) or "",
                "hint_notes": (h.notes if h else None) or "",
            }
        if mp == "wb" and it.wb_product:
            p = it.wb_product
            return p.id, {
                "barcode": p.barcode_primary or "",
                "name": p.name or "",
                "article": p.article or str(p.nm_id),
                "hint_packaging": "",
                "hint_notes": "",
            }
        return None

    for it in req.items:
        info = _product_key_and_meta(it)
        if info is None:
            unmatched.append({
                "marketplace": it.marketplace,
                "cluster": it.cluster,
                "raw_article": it.raw_article,
                "qty": it.qty,
            })
            continue
        key_id, meta = info
        slot = _fmt_slot(it.booked_slot_at) if it.booked_slot_at else None
        col_key: ColKey = (
            it.cluster,
            _clean_warehouse_name(it.target_warehouse),
            it.booked_supply_id or None,
            slot,
        )
        k = (col_key, key_id)
        by_col_sku[k] = by_col_sku.get(k, 0) + it.qty
        if key_id not in skus_meta:
            skus_meta[key_id] = meta

    wb_cols = _build_columns(req, "wb")
    oz_cols = _build_columns(req, "ozon")

    # Маппинг booked_supply_id (Ozon order_id, числовой) → метаданные поставки:
    # - order_number: номер заявки как в ЛК Ozon ('2000052363018'). fallback на supply_id.
    # - dropoff: финальная точка отгрузки (хаб для CROSSDOCK, РФЦ для DIRECT).
    # - slot: «17.05 13:00–14:00» из booked_slot_at.
    order_numbers: Dict[str, str] = {}
    dropoffs: Dict[str, str] = {}
    slots_full: Dict[str, str] = {}
    for it in req.items:
        if it.marketplace != "ozon" or not it.booked_supply_id:
            continue
        bsid = it.booked_supply_id
        if it.ozon_order_number:
            order_numbers[bsid] = it.ozon_order_number
        if it.ozon_dropoff_name:
            dropoffs[bsid] = it.ozon_dropoff_name
        if it.booked_slot_at:
            dt = it.booked_slot_at
            hh = dt.hour
            slots_full[bsid] = f"{dt.day:02d}.{dt.month:02d} {hh:02d}:00–{(hh + 1) % 24:02d}:00"

    def skus_for(mp: str) -> List[int]:
        seen: List[int] = []
        for it in req.items:
            info = _product_key_and_meta(it)
            if not info:
                continue
            if (it.marketplace or "").lower() == mp and info[0] not in seen:
                seen.append(info[0])
        return seen

    wb_skus = skus_for("wb")
    oz_skus = skus_for("ozon")

    wb = Workbook()
    wb.remove(wb.active)

    if wb_skus:
        ws = wb.create_sheet("вб")
        _fill_sheet(ws, "вб", wb_cols, wb_skus, skus_meta, by_col_sku,
                    order_numbers, dropoffs, slots_full)

    if oz_skus:
        ws = wb.create_sheet("озон")
        _fill_sheet(ws, "озон", oz_cols, oz_skus, skus_meta, by_col_sku,
                    order_numbers, dropoffs, slots_full)

    _fill_operations_sheet(wb.create_sheet("операции"), req, unmatched)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _fill_sheet(
    ws,
    kind: str,                                 # "вб" | "озон"
    cols: List[ColKey],
    sku_ids: List[int],
    skus_meta: Dict[int, Dict],
    by_col_sku: Dict[Tuple[ColKey, int], int],
    order_numbers: Optional[Dict[str, str]] = None,
    dropoffs: Optional[Dict[str, str]] = None,
    slots_full: Optional[Dict[str, str]] = None,
) -> None:
    header_fill = _header_fill()
    border = _thin_border()
    bold = Font(bold=True, name="Arial", size=11)
    normal = Font(name="Arial", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    is_oz = kind == "озон"
    # Ozon: «Артикул Поставщика» вынесен в самое начало (колонка A) — так удобнее
    # ФФ-сотруднику, артикул всегда виден первым. После него ШК / Название / Поставщик.
    if is_oz:
        fixed_left = ["Артикул Поставщика", "ШК", "Название товара", "Поставщик"]
        # «Полный адрес Озон склада» — только адрес drop-off. Таймслот теперь
        # отдельной строкой R3 под supply_id в колонке кластера.
        fixed_right = [
            "Упаковка для товара", "примечание",
            "Полный адрес Озон склада",
        ]
    else:
        fixed_left = ["ШК", "Название товара", "Поставщик"]
        fixed_right = ["упаковка", "состав набора", "Артикул Поставщика"]

    headers = fixed_left + [_col_label(c) for c in cols] + fixed_right
    n_cols = len(headers)

    # Шапка (R1)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 32

    # Шапка под R1 (одинаково для WB и Ozon):
    #   R2 = supply_id (под колонкой склада/кластера).
    #   R3 = дата+таймслот.
    # Drop-off для Ozon идёт в правую колонку «Полный адрес Озон склада» per-row.
    order_numbers = order_numbers or {}
    dropoffs = dropoffs or {}
    slots_full = slots_full or {}

    if is_oz:
        any_oz_book = any(c[2] for c in cols)
        if any_oz_book:
            first_dyn = len(fixed_left) + 1
            for i, col in enumerate(cols):
                col_idx = first_dyn + i
                _cluster, _wh, supply_id, _slot = col
                bsid = supply_id or ""
                num = order_numbers.get(bsid, bsid) if bsid else ""
                slot_s = slots_full.get(bsid, "") if bsid else ""
                r2 = ws.cell(row=2, column=col_idx, value=num)
                r3 = ws.cell(row=3, column=col_idx, value=slot_s)
                for r in (r2, r3):
                    r.font = normal
                    r.alignment = center
                    r.border = border
            # Заполнить остальные клетки строк 2/3 рамкой и пустотой.
            for c in range(1, n_cols + 1):
                for r in (2, 3):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border
                    if cell.value is None:
                        cell.value = ""
                    if not cell.font or cell.font.name != "Arial":
                        cell.font = normal
    else:
        first_dyn = len(fixed_left) + 1
        for i, col in enumerate(cols):
            col_idx = first_dyn + i
            _cluster, _wh, supply_id, slot = col
            r2 = ws.cell(row=2, column=col_idx, value=supply_id or "")
            r3 = ws.cell(row=3, column=col_idx, value=slot or "")
            for r in (r2, r3):
                r.font = normal
                r.alignment = center
                r.border = border
        for c in range(1, n_cols + 1):
            for r in (2, 3):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                if cell.value is None:
                    cell.value = ""
                if not cell.font or cell.font.name != "Arial":
                    cell.font = normal

    # Данные. Для Ozon: если есть бронирования — данные с R4 (как WB),
    # иначе с R2 (нет шапки supply_id/slot).
    if is_oz:
        any_oz_book = any(c[2] for c in cols)
        data_row_start = 4 if any_oz_book else 2
    else:
        data_row_start = 4
    row = data_row_start
    for sku_id in sku_ids:
        meta = skus_meta[sku_id]
        if is_oz:
            # Ozon: A=артикул, B=ШК, C=название, D=поставщик.
            ws.cell(row=row, column=1, value=meta["article"]).alignment = left
            ws.cell(row=row, column=2, value=meta["barcode"]).alignment = left
            ws.cell(row=row, column=3, value=meta["name"]).alignment = left
            ws.cell(row=row, column=4, value=DEFAULT_SUPPLIER).alignment = left
        else:
            # WB: A=ШК, B=название, C=поставщик (старый порядок).
            ws.cell(row=row, column=1, value=meta["barcode"]).alignment = left
            ws.cell(row=row, column=2, value=meta["name"]).alignment = left
            ws.cell(row=row, column=3, value=DEFAULT_SUPPLIER).alignment = left

        # qty per колонка
        first_dyn = len(fixed_left) + 1
        for i, col in enumerate(cols):
            qty = by_col_sku.get((col, sku_id))
            if qty:
                cell = ws.cell(row=row, column=first_dyn + i, value=qty)
                cell.alignment = center

        right_start = len(fixed_left) + len(cols) + 1
        # right block для WB: упаковка | состав набора | Артикул Поставщика
        # для Ozon: Упаковка | примечание | Полный адрес Озон + таймслот
        hint_pack = meta.get("hint_packaging") or None
        hint_notes = meta.get("hint_notes") or None
        c_pack = ws.cell(row=row, column=right_start, value=hint_pack)
        if hint_pack:
            c_pack.alignment = left
        c_notes = ws.cell(row=row, column=right_start + 1, value=hint_notes)
        if hint_notes:
            c_notes.alignment = left
        if is_oz:
            # Полный адрес drop-off: для каждого supply_id, в который у строки
            # есть qty, добавляем drop-off. Таймслот уже в R3 под supply_id.
            parts = []
            for col in cols:
                qty = by_col_sku.get((col, sku_id))
                if not qty:
                    continue
                _cluster, _wh, supply_id, _slot = col
                bsid = supply_id or ""
                drop = dropoffs.get(bsid, "")
                if drop:
                    parts.append(drop)
            if parts:
                ws.cell(row=row, column=right_start + 2,
                        value="; ".join(parts)).alignment = left
        else:
            # WB: 3-я колонка справа = Артикул Поставщика (как было).
            ws.cell(row=row, column=right_start + 2, value=meta["article"]).alignment = left

        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = border
            if not cell.font or cell.font.name != "Arial":
                cell.font = normal
        row += 1

    # Ширины колонок (из эталона) + динамика
    _apply_widths(ws, kind, n_cols, len(cols))


def _apply_widths(ws, kind: str, n_cols: int, n_dyn: int) -> None:
    """Подобраны под эталон пользователя (округлены)."""
    if kind == "вб":
        # A=19, B=28, C=15, dynamic=16, упаковка=28, состав=29, артикул=24
        ws.column_dimensions["A"].width = 19
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 15
        # dynamic columns
        dyn_start = 4
        for i in range(n_dyn):
            ws.column_dimensions[get_column_letter(dyn_start + i)].width = 16
        right_start = dyn_start + n_dyn
        widths_right = [28, 29, 24]
        for i, w in enumerate(widths_right):
            ws.column_dimensions[get_column_letter(right_start + i)].width = w
    else:
        # Ozon (новый порядок): A=Артикул(32), B=ШК(22), C=Название(32), D=Поставщик(14),
        # E…=кластеры(17), затем упаковка(24), примечание(29), Полный адрес(52).
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 32
        ws.column_dimensions["D"].width = 14
        dyn_start = 5
        for i in range(n_dyn):
            ws.column_dimensions[get_column_letter(dyn_start + i)].width = 17
        right_start = dyn_start + n_dyn
        widths_right = [24, 29, 52]
        for i, w in enumerate(widths_right):
            ws.column_dimensions[get_column_letter(right_start + i)].width = w


def _fill_operations_sheet(ws, req: ShipmentRequest, unmatched: List[Dict]) -> None:
    """Лист 'операции': metadata заявки + место для ручных заметок (merged A:B)."""
    border = _thin_border()
    bold = Font(bold=True, name="Arial", size=11)
    normal = Font(name="Arial", size=11)
    header_fill = _header_fill()
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    rows: List[Tuple[str, str]] = []
    rows.append(("Заявка", f"#{req.id}"))
    rows.append(("Создана", req.created_at.strftime("%Y-%m-%d %H:%M")))
    if req.target_date_from:
        date_s = req.target_date_from.strftime("%Y-%m-%d")
        if req.target_date_to:
            date_s += f" — {req.target_date_to:%Y-%m-%d}"
        rows.append(("Целевые даты", date_s))
    if req.target_dates_json:
        rows.append(("Конкретные даты", ", ".join(req.target_dates_json)))

    for i, (k, v) in enumerate(rows, 1):
        c1 = ws.cell(row=i, column=1, value=k)
        c2 = ws.cell(row=i, column=2, value=v)
        c1.font = bold
        c2.font = normal
        c1.alignment = left
        c2.alignment = left
        c1.border = border
        c2.border = border

    cur_row = len(rows) + 2
    if unmatched:
        ws.cell(row=cur_row, column=1, value="⚠ Позиции без SKU в каталоге:").font = bold
        cur_row += 1
        for h_idx, h in enumerate(["МП", "Кластер", "Артикул", "Кол-во"], 1):
            cell = ws.cell(row=cur_row, column=h_idx, value=h)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        cur_row += 1
        for u in unmatched:
            ws.cell(row=cur_row, column=1, value=u["marketplace"].upper()).border = border
            ws.cell(row=cur_row, column=2, value=u["cluster"]).border = border
            ws.cell(row=cur_row, column=3, value=u["raw_article"]).border = border
            ws.cell(row=cur_row, column=4, value=u["qty"]).border = border
            cur_row += 1
        cur_row += 1

    # Блок «📋 Операции / заметки» — merged A:B, для ручных заметок (как в эталоне)
    title = ws.cell(row=cur_row, column=1, value="📋 Операции / заметки")
    title.font = bold
    title.fill = header_fill
    title.alignment = center
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=2)
    title.border = border
    cur_row += 1
    for _ in range(10):
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=2)
        c = ws.cell(row=cur_row, column=1)
        c.border = border
        c.alignment = left
        c.font = normal
        cur_row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50
