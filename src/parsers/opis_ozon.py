"""Парсер 'Опись для заливки в OZ - *.xlsx'.

Реальная структура (golden file: переписки/files/МСК Опись для заливки в OZ - 03909.xlsx):
  row 1: ШК товара | Артикул товара | Кол-во товаров | Зона размещения | ШК ГМ | Тип ГМ | Срок годности
  row 2+: данные
"""
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import List, Optional

import openpyxl

from src.parsers.opis_wb import _to_date  # переиспользуем


@dataclass
class OpisOzonItem:
    barcode: str         # 'ШК товара' (на практике article)
    article: str         # 'Артикул товара'
    qty: int             # 'Кол-во товаров'
    zone: Optional[str]
    box_label: str       # 'ШК ГМ' — у Озона грузоместо
    box_type: Optional[str]
    expiry: Optional[date]


REQUIRED_HEADERS = ("шк товара", "кол-во товаров", "шк гм")


def _norm(v) -> str:
    return str(v or "").strip().lower()


def parse_opis_ozon(path: str | Path) -> List[OpisOzonItem]:
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active

    headers = [_norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    idx_map = {}
    for i, h in enumerate(headers):
        for key in ("шк товара", "артикул товара", "кол-во товаров",
                    "зона размещения", "шк гм", "тип гм", "срок годности"):
            if h.startswith(key):
                idx_map.setdefault(key, i)

    missing = [k for k in REQUIRED_HEADERS if k not in idx_map]
    if missing:
        raise ValueError(
            f"opis_ozon: missing headers {missing}. Got: {headers}. "
            "Запустить llm_fallback для smap-pинга."
        )

    items: List[OpisOzonItem] = []
    for r in range(2, ws.max_row + 1):
        bc = ws.cell(r, idx_map["шк товара"] + 1).value
        if bc is None or str(bc).strip() == "":
            continue
        qty = ws.cell(r, idx_map["кол-во товаров"] + 1).value
        if qty is None:
            continue

        article_v = ws.cell(r, idx_map["артикул товара"] + 1).value if "артикул товара" in idx_map else None
        zone_v = ws.cell(r, idx_map["зона размещения"] + 1).value if "зона размещения" in idx_map else None
        box_v = ws.cell(r, idx_map["шк гм"] + 1).value
        box_type_v = ws.cell(r, idx_map["тип гм"] + 1).value if "тип гм" in idx_map else None
        exp_v = ws.cell(r, idx_map["срок годности"] + 1).value if "срок годности" in idx_map else None

        items.append(OpisOzonItem(
            barcode=str(bc).strip(),
            article=str(article_v).strip() if article_v else "",
            qty=int(qty),
            zone=str(zone_v).strip() if zone_v else None,
            box_label=str(box_v).strip() if box_v else "",
            box_type=str(box_type_v).strip() if box_type_v else None,
            expiry=_to_date(exp_v),
        ))

    return items
