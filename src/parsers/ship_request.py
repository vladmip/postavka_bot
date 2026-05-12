"""Парсер xlsx-выгрузок из сервиса планирования отгрузок.

Поддерживает два формата:

WB (заголовок R1 = ['Баркод', 'Количество']):
    Баркод | Количество
    MILK-CHOCOLATE | 113
    2049738086922 | 15
    ...
    Имя файла: <Кластер>_<YYYY-MM-DD>_<HH-MM-SS>.xlsx, например 'Центральный_2026-05-11_18-25-30.xlsx'

Ozon (заголовок R1 = ['артикул', 'имя (необязательно)', 'количество']):
    артикул | имя (необязательно) | количество
    3CHOC | None | 158
    ...
    Имя файла: 'Москва, МО и Дальние регионы_2026-05-11_18-24-57.xlsx'

В колонке «артикул/баркод» может быть либо артикул поставщика (например MILK-CHOCOLATE),
либо численный штрих-код (например 2049738086922) — нормализатор каталога должен уметь оба.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook

logger = logging.getLogger("parsers.ship_request")


@dataclass
class ShipItem:
    article_or_barcode: str
    qty: int
    name_hint: Optional[str] = None


@dataclass
class ShipFileParsed:
    marketplace: str          # 'wb' | 'ozon'
    cluster_name: str         # 'Центральный', 'Москва, МО и Дальние регионы', ...
    items: List[ShipItem] = field(default_factory=list)
    file_name: str = ""


# Заголовки определяют формат
_WB_HEADERS = {"баркод", "количество"}
_OZ_HEADERS = {"артикул", "количество"}


def _norm(s) -> str:
    return (str(s) if s is not None else "").strip().lower()


def _detect_marketplace(headers: list) -> Optional[str]:
    h = {_norm(c) for c in headers if c}
    if _WB_HEADERS.issubset(h):
        return "wb"
    if _OZ_HEADERS.issubset(h):
        return "ozon"
    return None


# Имя файла: '<кластер>_YYYY-MM-DD_HH-MM-SS.xlsx'
# Telegram при отправке заменяет пробелы/запятые/дефисы на '_', поэтому
# дата может быть YYYY-MM-DD или YYYY_MM_DD.
_DATE_PART = r"\d{4}[-_]\d{2}[-_]\d{2}"
_NAME_RE = re.compile(rf"^(.+?)_{_DATE_PART}[_ ]", re.UNICODE)


def _extract_cluster_from_name(file_name: str) -> str:
    """Извлечь и нормализовать имя кластера из имени файла.

    Telegram заменяет ', ' '-' и пробелы на '_'. Превращаем обратно в пробелы
    (примерно). Например:
        'Москва,_МО_и_Дальние_регионы' → 'Москва МО и Дальние регионы'
    """
    name = Path(file_name).stem  # без .xlsx
    m = _NAME_RE.match(name + "_")  # добавим _ чтобы regex точно нашёл
    cluster = m.group(1) if m else name
    # '_,_' → ', ' (восстановление запятой) — встречается в Ozon
    cluster = cluster.replace(",_", ", ").replace("_,", ", ")
    # одиночные '_' → пробел
    cluster = cluster.replace("_", " ").strip()
    return cluster


def parse_ship_file(path: Path, original_name: Optional[str] = None) -> ShipFileParsed:
    """Парсит xlsx с распределением отгрузки на кластер."""
    file_name = original_name or Path(path).name
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Файл пустой")

    headers = list(rows[0])
    mp = _detect_marketplace(headers)
    if not mp:
        raise ValueError(
            f"Не распознал формат. Заголовки {headers!r}. Ожидаю либо WB ('Баркод|Количество') "
            "либо Ozon ('артикул|имя|количество')."
        )

    # Найдём индексы нужных колонок
    h_lower = [_norm(c) for c in headers]
    if mp == "wb":
        idx_art = h_lower.index("баркод")
        idx_qty = h_lower.index("количество")
        idx_name = None
    else:
        idx_art = h_lower.index("артикул")
        idx_qty = h_lower.index("количество")
        idx_name = h_lower.index("имя (необязательно)") if "имя (необязательно)" in h_lower else None

    items: List[ShipItem] = []
    for r in rows[1:]:
        art = r[idx_art]
        qty_raw = r[idx_qty]
        if art is None or qty_raw is None:
            continue
        art_s = str(art).strip()
        if not art_s:
            continue
        try:
            qty = int(float(qty_raw))
        except (ValueError, TypeError):
            logger.warning("Skip row with bad qty: %r", r)
            continue
        if qty <= 0:
            continue
        name_hint = None
        if idx_name is not None and r[idx_name]:
            name_hint = str(r[idx_name]).strip() or None
        items.append(ShipItem(article_or_barcode=art_s, qty=qty, name_hint=name_hint))

    return ShipFileParsed(
        marketplace=mp,
        cluster_name=_extract_cluster_from_name(file_name),
        items=items,
        file_name=file_name,
    )
