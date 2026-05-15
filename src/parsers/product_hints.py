"""Парсер xlsx с упаковкой/примечаниями для товаров Ozon.

Ожидаемый формат:
    артикул | упаковка | примечание
    3CHOC   | плёнка   | хрупкий
    ...

Названия колонок гибкие — ищем по подстрокам (case-insensitive, без пробелов).
Артикул — это offer_id на Ozon. Он же раскладывается в ozon_product_id на этапе сервиса.

Возвращает list[HintRow]: только строки с непустым артикулом. Пустые packaging/notes
сохраняются — на этапе upsert юзер сам решает, обнулять ли существующее значение.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook

logger = logging.getLogger("parsers.product_hints")


@dataclass
class HintRow:
    article: str
    packaging: Optional[str]
    notes: Optional[str]


_ARTICLE_ALIASES = ("артикул", "offer", "article")
_PACKAGING_ALIASES = ("упаков",)
_NOTES_ALIASES = ("примеч", "коммент", "notes", "note")


def _norm(s) -> str:
    return (str(s) if s is not None else "").strip().lower().replace(" ", "")


def _find_col(headers: list, aliases: tuple) -> Optional[int]:
    for idx, h in enumerate(headers):
        hn = _norm(h)
        if not hn:
            continue
        for a in aliases:
            if a in hn:
                return idx
    return None


def _cell_str(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def parse_hints_xlsx(file_path: str | Path) -> List[HintRow]:
    """Прочитать xlsx и вернуть список строк. Бросает ValueError если не найдена
    колонка артикула — без неё файл бесполезен."""
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        wb.close()
        return []

    art_idx = _find_col(headers, _ARTICLE_ALIASES)
    if art_idx is None:
        wb.close()
        raise ValueError(
            "Не найдена колонка артикула. Ожидаются колонки: артикул, упаковка, примечание."
        )
    pack_idx = _find_col(headers, _PACKAGING_ALIASES)
    notes_idx = _find_col(headers, _NOTES_ALIASES)

    out: List[HintRow] = []
    for raw in rows_iter:
        if not raw:
            continue
        if art_idx >= len(raw):
            continue
        art = _cell_str(raw[art_idx])
        if not art:
            continue
        packaging = _cell_str(raw[pack_idx]) if pack_idx is not None and pack_idx < len(raw) else None
        notes = _cell_str(raw[notes_idx]) if notes_idx is not None and notes_idx < len(raw) else None
        out.append(HintRow(article=art, packaging=packaging, notes=notes))

    wb.close()
    logger.info(
        "parse_hints_xlsx: %s → %d rows (pack_col=%s, notes_col=%s)",
        file_path, len(out), pack_idx, notes_idx,
    )
    return out
