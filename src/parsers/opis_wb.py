"""Парсер 'Опись для заливки в WB - *.xlsx'.

Реальная структура (golden file: переписки/files/Опись для заливки в WB - 03907.xlsx):
  row 1: Баркод товара | Кол-во товаров | ШК короба | срок годности
  row 2+: данные
"""
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import List, Optional

import openpyxl


@dataclass
class OpisWbItem:
    barcode: str          # 'Баркод товара' — на практике это article ('5CHOC-CARAMEL', '3CHOC')
    qty: int              # 'Кол-во товаров'
    box_label: str        # 'ШК короба' (например 'LBR_715921840073')
    expiry: Optional[date]  # 'срок годности'


REQUIRED_HEADERS = ("баркод товара", "кол-во товаров", "шк короба")


def _norm(v) -> str:
    return str(v or "").strip().lower()


def parse_opis_wb(path: str | Path) -> List[OpisWbItem]:
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active

    headers = [_norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers)}

    missing = [h for h in REQUIRED_HEADERS if h not in idx]
    if missing:
        raise ValueError(
            f"opis_wb: missing required headers {missing}. Got: {headers}. "
            "Запустить llm_fallback для smap-pинга."
        )

    barcode_i = idx["баркод товара"]
    qty_i = idx["кол-во товаров"]
    box_i = idx["шк короба"]
    expiry_i = idx.get("срок годности")

    items: List[OpisWbItem] = []
    for r in range(2, ws.max_row + 1):
        bc = ws.cell(r, barcode_i + 1).value
        if bc is None or str(bc).strip() == "":
            continue
        qty = ws.cell(r, qty_i + 1).value
        if qty is None:
            continue
        box = ws.cell(r, box_i + 1).value
        exp_raw = ws.cell(r, expiry_i + 1).value if expiry_i is not None else None
        exp = _to_date(exp_raw)

        items.append(OpisWbItem(
            barcode=str(bc).strip(),
            qty=int(qty),
            box_label=str(box).strip() if box else "",
            expiry=exp,
        ))

    return items


def _to_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    if hasattr(v, "date"):
        return v.date()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            from datetime import datetime as dt
            return dt.strptime(s, fmt).date()
        except ValueError:
            continue
    return None
