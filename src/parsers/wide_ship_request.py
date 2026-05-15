"""Парсер «широкого» xlsx-формата для пакетной заявки.

Структура входа:
    артикул | название | <Кластер 1> | <Кластер 2> | ... | <Кластер N>
    OFFER1  | …        | 5           |             | 3   |
    OFFER2  | …        |             | 2           |     |
    ...

Возвращает список ShipFileParsed: по одному «виртуальному файлу» на cluster.
Это позволяет переиспользовать существующий attach_ship_file для каждого cluster.

Детектирование широкого формата: есть колонка «артикул» И не выглядит как Ozon-узкий
формат (где есть колонка «количество» рядом).
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook

from src.parsers.ship_request import ShipItem, ShipFileParsed

logger = logging.getLogger("parsers.wide_ship_request")


def _norm(s) -> str:
    return (str(s) if s is not None else "").strip().lower()


# Известные кластеры Ozon FBO — используются для детектирования (если ≥2 совпадений
# в шапке → это широкий формат). Список взят из реальных файлов основной.xlsx /
# export(26).xlsx (Ozon ЛК → агрегированный экспорт остатков).
_KNOWN_CLUSTERS = {
    "москва, мо и дальние регионы", "санкт-петербург и сзо", "ростов", "новосибирск",
    "казань", "воронеж", "дальний восток", "екатеринбург", "ярославль", "красноярск",
    "уфа", "краснодар", "тюмень", "самара", "саратов", "невинномысск", "пермь",
    "махачкала", "омск", "беларусь", "калининград", "тверь", "оренбург", "астана", "алматы",
}


def _normalize_cluster_for_match(s: str) -> str:
    return s.strip().lower().replace("ё", "е")


def is_wide_format(headers: list) -> bool:
    """True если xlsx — широкий формат «артикул × кластеры». False если узкий.

    Поддерживает 2 варианта:
    1. Шаблон бота: A1='артикул', B1='название', C1+=кластеры.
    2. Экспорт из Ozon ЛК: A1=пусто, B1+=кластеры (артикул в A, начиная с R3).
    """
    h_norm = [_norm(c) for c in headers]
    # Узкий Ozon (артикул|имя|количество) или узкий WB (баркод|количество).
    if "количество" in h_norm:
        return False
    if "баркод" in h_norm:
        return False
    # Считаем матч по known cluster names в шапке.
    known_hits = sum(
        1 for h in h_norm
        if h and _normalize_cluster_for_match(h) in _KNOWN_CLUSTERS
    )
    if known_hits >= 2:
        return True
    # Fallback: есть явный заголовок «артикул» + ≥2 колонок справа.
    if "артикул" in h_norm and sum(1 for h in h_norm if h) >= 3:
        return True
    return False


def parse_wide_ship_file(
    path: Path,
    original_name: Optional[str] = None,
) -> List[ShipFileParsed]:
    """Парсит широкий xlsx и возвращает список ShipFileParsed (по cluster)."""
    file_name = original_name or Path(path).name
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Файл пустой")

    headers = list(rows[0])
    h_norm = [_norm(c) for c in headers]
    # Артикул — либо явный заголовок «артикул», либо колонка A (если её заголовок
    # пустой как в экспорте Ozon).
    if "артикул" in h_norm:
        idx_art = h_norm.index("артикул")
    elif not h_norm or h_norm[0] == "":
        idx_art = 0
    else:
        raise ValueError("Не найдена колонка артикула (ни заголовок «артикул», ни пустая A1)")
    idx_name = h_norm.index("название") if "название" in h_norm else None

    # Cluster-колонки: всё что не артикул/название и имеет непустой заголовок.
    cluster_cols: List[tuple] = []  # [(col_idx, cluster_name)]
    for i, h in enumerate(headers):
        if i == idx_art or i == idx_name:
            continue
        if h is None:
            continue
        cluster_name = str(h).strip()
        if not cluster_name:
            continue
        cluster_cols.append((i, cluster_name))

    if not cluster_cols:
        raise ValueError("Не найдено ни одной колонки-кластера")

    # Валидация имён кластеров — должны совпадать с known Ozon-кластерами.
    unknown = [
        name for _, name in cluster_cols
        if _normalize_cluster_for_match(name) not in _KNOWN_CLUSTERS
    ]
    if unknown:
        raise ValueError(
            "Неизвестные кластеры в шапке файла:\n  • "
            + "\n  • ".join(unknown)
            + "\n\nДопустимые: " + ", ".join(sorted(_KNOWN_CLUSTERS))
        )

    by_cluster: dict = {name: [] for _, name in cluster_cols}
    skipped = 0
    for r in rows[1:]:
        if not r:
            continue
        art = r[idx_art] if idx_art < len(r) else None
        if art is None:
            # Возможна служебная строка с totals (Ozon экспорт R2) — пропускаем.
            continue
        art_s = str(art).strip()
        if not art_s:
            continue
        name_hint = None
        if idx_name is not None and idx_name < len(r) and r[idx_name]:
            name_hint = str(r[idx_name]).strip() or None
        for col_idx, cluster_name in cluster_cols:
            if col_idx >= len(r):
                continue
            qty_raw = r[col_idx]
            if qty_raw is None or qty_raw == "":
                continue
            try:
                qty = int(float(qty_raw))
            except (ValueError, TypeError):
                skipped += 1
                continue
            if qty <= 0:
                continue
            by_cluster[cluster_name].append(
                ShipItem(article_or_barcode=art_s, qty=qty, name_hint=name_hint)
            )

    if skipped:
        logger.info("parse_wide_ship_file: %d cells with bad qty skipped", skipped)

    out: List[ShipFileParsed] = []
    for _, cluster_name in cluster_cols:
        items = by_cluster.get(cluster_name) or []
        if not items:
            continue
        out.append(ShipFileParsed(
            marketplace="ozon",
            cluster_name=cluster_name,
            items=items,
            file_name=file_name,
        ))
    if not out:
        raise ValueError("Не нашлось ни одной заполненной строки с qty > 0")
    return out
