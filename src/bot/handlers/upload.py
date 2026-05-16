"""Обработка пересланных пользователем xlsx/xls от ФФ.

Workflow:
1. Скачать файл во `data/storage/`
2. Классифицировать (router.classify_file)
3. Распарсить
4. Если parsed — спросить «к какой поставке привязать?»
5. После привязки — обновить supply_items (для опись_*) или показать расхождения (для prihod)
"""
import json
from datetime import datetime
from pathlib import Path

from aiogram import Router, F
from aiogram.types import Message, CallbackQuery
from aiogram.fsm.context import FSMContext

from src.bot.helpers import safe_edit_or_answer
from src.bot.keyboards import kb_pick_supply
from src.bot.states import UploadBind
from src.config import STORAGE_DIR
from src.db.session import db_session
from src.db.models import InboxFile
from src.parsers import (
    classify_file, FileKind,
    parse_opis_wb, parse_opis_ozon, parse_prihod, parse_ostatki,
)
from src.services.supply_service import (
    list_supplies, attach_picked_qty, transition,
)
from src.services.reconciler import reconcile_prihod

router = Router()


_SCREENSHOTS_DIR = Path(__file__).parent.parent.parent.parent / "data" / "screenshots"
_SCREENSHOTS_DIR.mkdir(parents=True, exist_ok=True)


@router.message(F.photo)
async def handle_photo(msg: Message) -> None:
    """Сохраняем скрины в data/screenshots/ — чтобы Claude мог их Read'ом читать."""
    photo = msg.photo[-1]  # самая большая версия
    ts = datetime.utcnow().strftime("%Y-%m-%d_%H-%M-%S")
    fname = f"{ts}_{photo.file_unique_id}.jpg"
    target = _SCREENSHOTS_DIR / fname
    file = await msg.bot.get_file(photo.file_id)
    await msg.bot.download_file(file.file_path, destination=target)
    caption_extra = f"\n💬 Подпись: <i>{msg.caption}</i>" if msg.caption else ""
    await msg.answer(
        f"📷 Скрин сохранён:\n<code>{target.as_posix()}</code>{caption_extra}\n\n"
        f"<i>Скажи мне в чате — я открою и посмотрю что на нём.</i>"
    )


# Защита: не пытаемся обрабатывать файлы > 20 МБ. Telegram сам ограничивает
# upload до 50 МБ для ботов, но xlsx больше 20 — это либо что-то экзотическое,
# либо кто-то пытается положить процесс. Обычные выгрузки Ozon/WB < 5 МБ.
_MAX_DOC_SIZE_MB = 20


@router.message(F.document)
async def handle_document(msg: Message, state: FSMContext) -> None:
    doc = msg.document
    fname = doc.file_name or "upload.bin"

    # Защита от больших файлов (DoS / accidental упор в RAM на parsing).
    if doc.file_size and doc.file_size > _MAX_DOC_SIZE_MB * 1024 * 1024:
        await msg.answer(
            f"⚠ Файл слишком большой ({doc.file_size // (1024*1024)} МБ). "
            f"Лимит: {_MAX_DOC_SIZE_MB} МБ. Обычные xlsx-выгрузки сильно меньше — проверь файл."
        )
        return

    # Zip с xlsx-выгрузками — распаковываем и обрабатываем каждый xlsx по очереди.
    if fname.lower().endswith(".zip"):
        await _handle_zip(msg, state, doc, fname)
        return

    if not (fname.lower().endswith(".xls") or fname.lower().endswith(".xlsx")):
        await msg.answer("Принимаю только .xls / .xlsx или .zip с ними внутри.")
        return

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    stored_path = STORAGE_DIR / f"{ts}_{fname}"

    file = await msg.bot.get_file(doc.file_id)
    await msg.bot.download_file(file.file_path, destination=stored_path)

    # Перед стандартным sheep_handler — попробуем широкий формат «артикул × кластеры».
    # Это шаблон «➕ Новая заявка» — заполненный xlsx с количеством per cluster.
    try:
        from openpyxl import load_workbook
        from src.parsers.wide_ship_request import is_wide_format, parse_wide_ship_file
        wb = load_workbook(stored_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        first_row = next(ws.iter_rows(values_only=True), None) or []
        wb.close()
        if is_wide_format(list(first_row)):
            await _handle_wide_ship_file(msg, state, stored_path, fname)
            return
    except Exception as e:
        logger.warning("wide format probe failed for %s: %s", fname, e)

    # Если имя файла похоже на ship-выгрузку → отдаём shipment handler
    from src.bot.handlers.shipment import looks_like_ship_file, handle_ship_document
    if looks_like_ship_file(fname):
        await handle_ship_document(msg, state, stored_path, fname)
        return

    kind = classify_file(stored_path)

    if kind == FileKind.UNKNOWN:
        await msg.answer(f"Не распознал тип файла: {fname}. Сохранил в storage/.")
        _save_inbox(stored_path, fname, kind.value, None, {})
        return

    try:
        parsed_summary, parsed_payload = _parse(stored_path, kind)
    except Exception as e:
        await msg.answer(f"⚠ Парсер упал на {fname}: {e}")
        _save_inbox(stored_path, fname, kind.value, None, {"error": str(e)})
        return

    inbox_id = _save_inbox(stored_path, fname, kind.value, None, parsed_payload)

    with db_session() as session:
        candidates = list_supplies(session, limit=20)

    if not candidates:
        await msg.answer(
            f"✅ Распознал: <b>{kind.value}</b>\n{parsed_summary}\n\n"
            f"Поставок в БД нет — привязать не к чему. Сначала /supply_new."
        )
        return

    await state.update_data(inbox_id=inbox_id, kind=kind.value, stored_path=str(stored_path))
    await state.set_state(UploadBind.pick_supply)
    await msg.answer(
        f"✅ Распознал: <b>{kind.value}</b>\n{parsed_summary}\n\n"
        f"К какой поставке привязать?",
        reply_markup=kb_pick_supply(candidates),
    )


@router.callback_query(UploadBind.pick_supply, F.data.startswith("bind:"))
async def cb_bind(cb: CallbackQuery, state: FSMContext) -> None:
    supply_id = int(cb.data.split(":", 1)[1])
    data = await state.get_data()
    kind = data["kind"]
    stored_path = Path(data["stored_path"])
    inbox_id = data["inbox_id"]

    await state.clear()

    with db_session() as session:
        inbox = session.get(InboxFile, inbox_id)
        if inbox:
            inbox.supply_id = supply_id

        if kind == FileKind.OPIS_WB.value:
            items = parse_opis_wb(stored_path)
            opis_tuples = [(i.barcode, i.qty, i.box_label, i.expiry) for i in items]
            result = attach_picked_qty(session, supply_id, opis_tuples)
            transition(session, supply_id, "picked", event=f"opis_wb {stored_path.name}")
            await cb.answer("Привязано")
            if cb.message:
                miss = result["missing"]
                miss_txt = f"\nНе нашёл в поставке: {', '.join(miss)}" if miss else ""
                await safe_edit_or_answer(cb.message,
                    f"📎 Привязал опись WB к поставке #{supply_id}. "
                    f"Совпало: {result['matched']}/{len(opis_tuples)}.{miss_txt}\n"
                    f"Состояние → picked"
                )

        elif kind == FileKind.OPIS_OZON.value:
            items = parse_opis_ozon(stored_path)
            opis_tuples = [(i.barcode, i.qty, i.box_label, i.expiry) for i in items]
            result = attach_picked_qty(session, supply_id, opis_tuples)
            transition(session, supply_id, "picked", event=f"opis_ozon {stored_path.name}")
            await cb.answer("Привязано")
            if cb.message:
                miss = result["missing"]
                miss_txt = f"\nНе нашёл в поставке: {', '.join(miss)}" if miss else ""
                await safe_edit_or_answer(cb.message,
                    f"📎 Привязал опись Озон к поставке #{supply_id}. "
                    f"Совпало: {result['matched']}/{len(opis_tuples)}.{miss_txt}\n"
                    f"Состояние → picked"
                )

        elif kind == FileKind.PRIHOD.value:
            doc = parse_prihod(stored_path)
            prihod_tuples = [(i.article, i.qty) for i in doc.items]
            discrep = reconcile_prihod(session, supply_id, prihod_tuples)
            transition(session, supply_id, "intake_done",
                       event=f"prihod {doc.doc_number or stored_path.name}")
            lines = [
                f"📎 Приходная привязана к поставке #{supply_id}. Состояние → intake_done"
            ]
            if discrep:
                lines.append(f"\n⚠ Расхождений: {len(discrep)}")
                for d in discrep[:20]:
                    sign = "+" if d.delta > 0 else ""
                    lines.append(f"  <code>{d.article}</code>: план {d.expected}, факт {d.actual} ({sign}{d.delta})")
            else:
                lines.append("\n✅ Расхождений нет.")
            await cb.answer("Сверено")
            if cb.message:
                await safe_edit_or_answer(cb.message,"\n".join(lines))

        elif kind == FileKind.OSTATKI.value:
            items = parse_ostatki(stored_path)
            await cb.answer()
            lines = [f"📎 Остатки от ФФ привязал к контексту поставки #{supply_id}.",
                     f"Записей: {len(items)}"]
            top = sorted(items, key=lambda x: -x.balance)[:10]
            for i in top:
                lines.append(f"  <code>{i.article}</code>: {i.balance} (дост {i.available} / рез {i.reserved})")
            if cb.message:
                await safe_edit_or_answer(cb.message,"\n".join(lines))


def _parse(path: Path, kind: FileKind) -> tuple[str, dict]:
    if kind == FileKind.OPIS_WB:
        items = parse_opis_wb(path)
        summary = f"Опись WB: {len(items)} позиций, ШК короб {items[0].box_label if items else '?'}"
        payload = [
            {"barcode": i.barcode, "qty": i.qty, "box": i.box_label, "exp": str(i.expiry) if i.expiry else None}
            for i in items
        ]
        return summary, {"items": payload}

    if kind == FileKind.OPIS_OZON:
        items = parse_opis_ozon(path)
        summary = f"Опись Озон: {len(items)} позиций"
        payload = [
            {"barcode": i.barcode, "article": i.article, "qty": i.qty,
             "box": i.box_label, "exp": str(i.expiry) if i.expiry else None}
            for i in items
        ]
        return summary, {"items": payload}

    if kind == FileKind.PRIHOD:
        doc = parse_prihod(path)
        summary = f"Приходная № {doc.doc_number or '?'}: {len(doc.items)} позиций"
        payload = {
            "doc_number": doc.doc_number,
            "items": [{"article": i.article, "qty": i.qty, "total": i.total} for i in doc.items],
        }
        return summary, payload

    if kind == FileKind.OSTATKI:
        items = parse_ostatki(path)
        summary = f"Остатки: {len(items)} SKU"
        payload = {
            "items": [{"article": i.article, "balance": i.balance,
                       "available": i.available, "reserved": i.reserved}
                      for i in items]
        }
        return summary, payload

    return "?", {}


def _save_inbox(path: Path, original_name: str, kind: str, supply_id, payload: dict) -> int:
    with db_session() as session:
        inbox = InboxFile(
            original_name=original_name,
            file_kind=kind,
            supply_id=supply_id,
            parsed_payload_json=payload,
            file_path=str(path),
        )
        session.add(inbox)
        session.flush()
        return inbox.id


# ── Широкий xlsx (новая заявка одной таблицей) ───────────────────────────


async def _handle_wide_ship_file(msg, state: FSMContext, path: Path, fname: str) -> None:
    """Принять xlsx с широкой структурой «артикул × кластеры» — это шаблон
    «Новая заявка». Создаёт одну ShipmentRequest, приатачивает items per cluster."""
    from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup
    from src.parsers.wide_ship_request import parse_wide_ship_file
    from src.services.shipment_service import create_shipment_request, attach_ship_file
    logger = __import__("logging").getLogger("bot.upload")

    try:
        parsed_list = parse_wide_ship_file(path, original_name=fname)
    except ValueError as e:
        # Валидационные ошибки (неизвестные кластеры, нет артикулов) — показываем
        # как понятную ошибку без раскрытия трейса.
        await msg.answer(
            f"❌ <b>Файл не принят</b>\n\n{e}\n\n"
            "Поправь файл и пришли заново."
        )
        return
    except Exception as e:
        logger.exception("wide parse failed: %s", e)
        await msg.answer(
            f"❌ <b>Не распарсил xlsx</b>\n<code>{type(e).__name__}: {e}</code>"
        )
        return

    # Pre-validation: суммируем артикулы, проверяем их наличие в каталоге ДО создания.
    # Multi-tenant: ищем в каталоге конкретного юзера, не глобально.
    from src.services.shipment_service import _find_ozon_product
    from src.services.user_service import current_user_id_from
    tg_id = current_user_id_from(msg)
    all_articles = {it.article_or_barcode for p in parsed_list for it in p.items}
    with db_session() as session:
        unmatched_articles = sorted(
            art for art in all_articles
            if _find_ozon_product(session, art, user_id=tg_id) is None
        )
    if unmatched_articles:
        # Жёсткая остановка: половина и больше — отбиваем.
        sample = ", ".join(f"<code>{a}</code>" for a in unmatched_articles[:10])
        more = (
            f" … и ещё {len(unmatched_articles) - 10}"
            if len(unmatched_articles) > 10 else ""
        )
        if len(unmatched_articles) >= max(1, len(all_articles) // 2):
            await msg.answer(
                f"❌ <b>Слишком много неизвестных артикулов</b>\n\n"
                f"Не нашлось в Ozon-каталоге: <b>{len(unmatched_articles)}</b> из <b>{len(all_articles)}</b>:\n"
                f"  {sample}{more}\n\n"
                "Сначала обнови каталог: ⚙ Настройки → 🔗 Привязать каталог к МП."
            )
            return
        # Иначе — предупреждаем, но создаём заявку (matched попадут, unmatched в reject-список).
        await msg.answer(
            f"⚠ <b>Часть артикулов не найдена в каталоге ({len(unmatched_articles)} шт):</b>\n"
            f"  {sample}{more}\n\n"
            "Они попадут в «без SKU» и не будут отправлены в Ozon. Создаю заявку."
        )

    total_items = sum(len(p.items) for p in parsed_list)
    total_qty = sum(sum(it.qty for it in p.items) for p in parsed_list)
    clusters = [p.cluster_name for p in parsed_list]

    with db_session() as session:
        req = create_shipment_request(session, source_file=fname, user_id=tg_id)
        rid = req.id
        per_cluster = []
        for parsed in parsed_list:
            result = attach_ship_file(session, rid, parsed, user_id=tg_id)
            per_cluster.append((parsed.cluster_name, result))

    lines = [
        f"📦 <b>Создана заявка #{rid}</b>",
        f"Кластеров: <b>{len(clusters)}</b> · Позиций: <b>{total_items}</b> · Количество: <b>{total_qty}</b>",
        "",
        "<b>По кластерам:</b>",
    ]
    for cl, result in per_cluster:
        unm = f" ⚠ {len(result.unmatched_articles)} без SKU" if result.unmatched_articles else ""
        lines.append(f"  • {cl}: {result.matched} SKU{unm}")
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📋 Открыть поставку", callback_data=f"ship_open:{rid}")],
        [InlineKeyboardButton(text="📋 Все заявки", callback_data="menu:ships")],
    ])
    await msg.answer("\n".join(lines), reply_markup=kb)


# ── ZIP handling ─────────────────────────────────────────────────────────


async def _handle_zip(msg, state: FSMContext, doc, fname: str) -> None:
    """Принять zip-архив, распаковать и спросить юзера: одна заявка на все
    или каждый файл — отдельная заявка."""
    import zipfile
    from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    stored_zip = STORAGE_DIR / f"{ts}_{fname}"
    file = await msg.bot.get_file(doc.file_id)
    await msg.bot.download_file(file.file_path, destination=stored_zip)

    extract_dir = STORAGE_DIR / f"{ts}_extracted"
    extract_dir.mkdir(parents=True, exist_ok=True)

    try:
        with zipfile.ZipFile(stored_zip, "r") as zf:
            zf.extractall(extract_dir)
    except zipfile.BadZipFile:
        await msg.answer(f"⚠ Файл <code>{fname}</code> не валидный zip.")
        return

    xlsx_files = []
    for p in extract_dir.rglob("*"):
        if p.is_file() and p.suffix.lower() in (".xls", ".xlsx"):
            xlsx_files.append(p)

    if not xlsx_files:
        await msg.answer(f"⚠ В zip {fname} нет .xls/.xlsx файлов.")
        return

    # Сохраняем пути в state и спрашиваем как обрабатывать
    await state.update_data(zip_paths=[str(p) for p in xlsx_files], zip_name=fname)
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🚚 Одна заявка (вместе)", callback_data="zip_mode:together")],
        [InlineKeyboardButton(text="📤 Раздельно — N заявок", callback_data="zip_mode:separate")],
        [InlineKeyboardButton(text="✖ Отмена", callback_data="zip_mode:cancel")],
    ])
    names = ", ".join(f"<code>{p.name[:35]}</code>" for p in xlsx_files[:8])
    if len(xlsx_files) > 8:
        names += f", …и ещё {len(xlsx_files) - 8}"
    await msg.answer(
        f"📦 В zip <code>{fname}</code> — {len(xlsx_files)} xlsx:\n{names}\n\n"
        f"Как обрабатывать?",
        reply_markup=kb,
    )


@router.callback_query(F.data.startswith("zip_mode:"))
async def cb_zip_mode(cb: CallbackQuery, state: FSMContext) -> None:
    mode = cb.data.split(":", 1)[1]
    data = await state.get_data()
    paths = [Path(p) for p in (data.get("zip_paths") or [])]
    zip_name = data.get("zip_name") or "zip"
    await state.update_data(zip_paths=None, zip_name=None)

    if mode == "cancel" or not paths:
        await cb.answer("Отменено")
        return

    from src.bot.handlers.shipment import (
        looks_like_ship_file, handle_ship_document,
        _create_zip_together_request, _ask_ozon_type_for_new,
    )
    from src.parsers.ship_request import parse_ship_file

    await cb.answer("Обрабатываю…")
    if not cb.message:
        return

    logger = __import__("logging").getLogger("bot.upload")

    if mode == "together":
        # Все xlsx → ОДНА заявка. Если среди файлов есть хотя бы один Ozon —
        # сначала спрашиваем тип поставки (новый wizard-шаг), потом создаём
        # заявку и приатачиваем все файлы.
        ship_paths = [p for p in paths if looks_like_ship_file(p.name)]
        has_ozon = False
        for p in ship_paths:
            try:
                parsed = parse_ship_file(p)
            except Exception:
                continue
            if parsed.marketplace == "ozon":
                has_ozon = True
                break
        if has_ozon:
            await state.update_data(
                up_otype_kind="zip",
                up_otype_zip_paths=[str(p) for p in paths],
                up_otype_zip_name=zip_name,
            )
            await _ask_ozon_type_for_new(
                cb.message, state,
                header=f"📦 zip <code>{zip_name}</code> · {len(paths)} файлов",
            )
            return
        await _create_zip_together_request(cb.message, state, paths, zip_name, otype=None)

    elif mode == "separate":
        # Каждый xlsx → своя заявка. БЕЗ вопросов про тип Ozon — юзер выберет
        # потом в карточке. Цель: моментальная пакетная нарезка.
        from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup
        from src.services.shipment_service import create_shipment_request, attach_ship_file
        from src.db.session import db_session

        ok, errs = 0, 0
        created_rids: list = []
        for path in paths:
            inner_name = path.name
            if looks_like_ship_file(inner_name):
                try:
                    parsed = parse_ship_file(path, original_name=inner_name)
                    with db_session() as session:
                        req = create_shipment_request(session, source_file=inner_name)
                        attach_ship_file(session, req.id, parsed)
                        rid = req.id
                    ok += 1
                    created_rids.append((rid, parsed.marketplace, parsed.cluster_name, len(parsed.items)))
                except Exception as e:
                    errs += 1
                    logger.exception("zip separate: failed %s", inner_name)
                    await cb.message.answer(
                        f"⚠ <code>{inner_name}</code>: {type(e).__name__}: {str(e)[:200]}"
                    )
            else:
                try:
                    kind = classify_file(path)
                    if kind == FileKind.UNKNOWN:
                        await cb.message.answer(f"❔ <code>{inner_name}</code>: тип не определён.")
                        continue
                    parsed_summary, parsed_payload = _parse(path, kind)
                    _save_inbox(path, inner_name, kind.value, None, parsed_payload)
                    await cb.message.answer(f"✅ <code>{inner_name}</code>: {kind.value}\n{parsed_summary}")
                except Exception as e:
                    await cb.message.answer(f"⚠ <code>{inner_name}</code>: {type(e).__name__}: {str(e)[:200]}")

        if created_rids:
            mp_emoji = {"wb": "🟣", "ozon": "🔵"}
            lines = [f"📦 <b>Создано заявок: {ok}</b>"]
            kb_rows = []
            for rid, mp, cluster, n in created_rids:
                emoji = mp_emoji.get(mp, "•")
                lines.append(f"  {emoji} #{rid} — {cluster} ({n} SKU)")
                kb_rows.append([InlineKeyboardButton(
                    text=f"📋 Заявка #{rid} — {cluster[:30]}",
                    callback_data=f"ship_open:{rid}",
                )])
            kb_rows.append([InlineKeyboardButton(text="📋 Все заявки", callback_data="menu:ships")])
            if errs:
                lines.append(f"\n⚠ Ошибок: {errs}")
            await cb.message.answer(
                "\n".join(lines),
                reply_markup=InlineKeyboardMarkup(inline_keyboard=kb_rows),
            )
